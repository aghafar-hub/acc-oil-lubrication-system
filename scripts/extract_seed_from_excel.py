import openpyxl, json, re, hashlib
from datetime import datetime, date
from collections import defaultdict, Counter

wb = openpyxl.load_workbook('/mnt/user-data/uploads/ACC_Lubricants_Master_v2.xlsx', data_only=True)
M = wb['Master_LP_DB']; CL = wb['LP_Change LOG']; LPS = wb['Lubrication Points']
mh = [c.value for c in M[1]]; ch = [c.value for c in CL[1]]; lh = [c.value for c in LPS[1]]

def mv(r, col): return M.cell(row=r, column=mh.index(col)+1).value
def cv(r, col): return CL.cell(row=r, column=ch.index(col)+1).value
def lv(r, col): return LPS.cell(row=r, column=lh.index(col)+1).value
def s(v):
    if v is None: return None
    v = str(v).strip()
    return v if v else None

CAL_DAYS = {'0.5 Y':183,'1 Y':365,'1.5 Y':548,'2 Y':730,'3 Y':1095,'4 Y':1460,'5 Y':1825}
MONTHS = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Sept':9,'Oct':10,'Nov':11,'Dec':12}

def parse_month_year(v):
    """Parse 'Apr 2026' style strings to an ISO date (day=1)."""
    v = s(v)
    if not v: return None
    if isinstance(v, datetime): return v.date().isoformat()
    m = re.match(r'([A-Za-z]+)\.?\s+(\d{4})', v)
    if m:
        mon = MONTHS.get(m.group(1)[:3].capitalize() if len(m.group(1))>3 else m.group(1).capitalize())
        if not mon:
            mon = MONTHS.get(m.group(1).capitalize())
        if mon:
            return date(int(m.group(2)), mon, 1).isoformat()
    return None

def to_iso(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, date): return v.isoformat()
    return None

data_quality_flags = []

def safe_numeric(v, lp_id_code, field_name):
    """Parse a numeric field robustly; if it contains non-numeric text
    (source data-entry error), return None and log it for client review
    instead of silently guessing or crashing the import."""
    if v is None: return None
    if isinstance(v, (int, float)): return v
    txt = str(v).replace('\xa0', ' ').replace(',', '').strip()
    if not txt: return None
    try:
        return float(txt)
    except ValueError:
        data_quality_flags.append({
            "lp_id_code": lp_id_code, "field": field_name, "raw_value": str(v),
            "action_taken": "imported as null — needs correction in source data"
        })
        return None

# ---------- Organizations ----------
organizations = [
    {"id": "org-acc", "name": "ACC", "type": "owner"},
    {"id": "org-rhi", "name": "RHI", "type": "contractor"},
    {"id": "org-asec", "name": "ASEC", "type": "contractor"},
]
org_by_name = {"ACC": "org-acc", "RHI": "org-rhi", "ASEC": "org-asec"}

# ---------- Areas (from Locn. code, contractor verified 1:1 consistent) ----------
locn_org = {}
locn_count = Counter()
for r in range(2, M.max_row+1):
    locn = mv(r, 'Locn.')
    if locn is None: continue
    contractor = s(mv(r, 'Contractor '))
    locn_org[locn] = org_by_name.get(contractor, "org-acc")
    locn_count[locn]+=1

areas = []
for locn in sorted(locn_org.keys()):
    areas.append({
        "id": f"area-{locn}",
        "locn_code": str(locn),
        "name": f"Area {locn}",
        "organization_id": locn_org[locn],
    })

# ---------- Equipment ----------
equip_rows = defaultdict(list)
for r in range(2, M.max_row+1):
    asset = s(mv(r, 'Asset'))
    if not asset: continue
    equip_rows[asset].append(r)

equipment = []
eq_id_map = {}
for asset, rows in equip_rows.items():
    r0 = rows[0]
    locn = mv(r0, 'Locn.')
    asset_name = s(mv(r0, 'Asset Name'))
    gb = None; temp = None; rh = None
    for rr in rows:
        gb = gb or s(mv(rr, 'Gear box brand'))
        temp = temp if temp is not None else mv(rr, 'Op. Temp. ◦C')
        rh = rh if rh is not None else mv(rr, 'Annual RH Actual')
    eid = f"eq-{asset}"
    eq_id_map[asset] = eid
    equipment.append({
        "id": eid,
        "equipment_id_code": asset,
        "asset_name": asset_name,
        "area_id": f"area-{locn}" if locn is not None else None,
        "gearbox_brand": gb,
        "op_temp_c": safe_numeric(temp, asset, "Op. Temp. ◦C"),
        "annual_rh_actual": safe_numeric(rh, asset, "Annual RH Actual"),
    })

# ---------- Lubricant Types ----------
lt_map = {}
lubricant_types = []
for r in range(2, M.max_row+1):
    name = s(mv(r, 'Lubricant Type'))
    brand = s(mv(r, 'Lubricant brand'))
    if not name: continue
    key = (name, brand)
    if key not in lt_map:
        ltid = f"lt-{len(lubricant_types)+1}"
        lt_map[key] = ltid
        lubricant_types.append({"id": ltid, "name": name, "brand": brand})

def freq_bucket(freq):
    freq = s(freq)
    if freq in CAL_DAYS: return "calendar", CAL_DAYS[freq], freq
    if freq == "Oil Analysis": return "oil_analysis", None, freq
    if freq in ("As needed", "If needed"): return "as_needed", None, freq
    return "unknown", None, freq

# ---------- Lubrication Points (902 rows, aligned across sheets by row index) ----------
lubrication_points = []
seen_lp_ids = set()
skipped_no_lpid = 0
for idx, r in enumerate(range(2, M.max_row+1)):
    asset = s(mv(r, 'Asset'))
    if not asset:
        continue
    lp_id_code = s(lv(r, 'Lubrication_ID'))  # cleaned unique ID from Lubrication Points sheet
    if not lp_id_code:
        skipped_no_lpid += 1
        continue
    if lp_id_code in seen_lp_ids:
        # extremely defensive — shouldn't happen, Lubrication_ID was verified unique
        lp_id_code = f"{lp_id_code}-DUP{idx}"
    seen_lp_ids.add(lp_id_code)

    freq_raw = mv(r, 'Frequncy')
    ftype, fdays, freq_label = freq_bucket(freq_raw)

    name = s(mv(r, 'Lubricant Type')); brand = s(mv(r, 'Lubricant brand'))
    ltid = lt_map.get((name, brand)) if name else None

    lubrication_points.append({
        "id": f"lp-{idx+1}",
        "lp_id_code": lp_id_code,
        "equipment_id": eq_id_map.get(asset),
        "point_description": s(lv(r, 'Lubrication_Point_Raw')) or s(mv(r, 'Lubrication Points')),
        "point_code": s(lv(r, 'Point_Code')),
        "position": s(lv(r, 'Position')),
        "standard_quantity_l": safe_numeric(mv(r, 'Qty (ltr.)'), lp_id_code, "Qty (ltr.)"),
        "lubricant_type_id": ltid,
        "frequency_type": ftype,
        "frequency_label": freq_label,
        "frequency_interval_days": fdays,
        "oh_hours_reference": safe_numeric(mv(r, 'OH (hrs.)'), lp_id_code, "OH (hrs.)"),
        "oa_required": s(mv(r, 'OA Required')) == "Yes",
        "oa_last_sample_date": parse_month_year(mv(r, 'OA Last Sample')),
        "oa_interval_days": safe_numeric(mv(r, 'OA Interval Days'), lp_id_code, "OA Interval Days"),
        "oa_interval_label": s(mv(r, 'OA Interval')),
        "remarks": s(mv(r, 'Remarks')),
        "_row": r,  # internal, used to join change-log below; stripped before final write
    })

print("Lubrication points extracted:", len(lubrication_points), "skipped (no LP ID):", skipped_no_lpid)
print("Frequency type counts:", Counter(p['frequency_type'] for p in lubrication_points))

# ---------- Historical lubrication_records from LP_Change LOG ----------
CHANGE_COLS = [f"Change Date {i}" for i in range(1, 13)]
lubrication_records = []
rec_seq = 1
for p in lubrication_points:
    r = p["_row"]
    dates = []
    for col in CHANGE_COLS:
        v = cv(r, col)
        iso = to_iso(v)
        if iso: dates.append(iso)
    last_change = to_iso(cv(r, 'Last Change Date'))
    if last_change and last_change not in dates:
        dates.append(last_change)
    dates = sorted(set(dates))
    for d in dates:
        lubrication_records.append({
            "id": f"rec-{rec_seq}",
            "lp_id": p["id"],
            "technician_id": None,  # legacy import, no technician of record
            "lubrication_date": d,
            "quantity_used_l": p["standard_quantity_l"],
            "oil_type_used_id": p["lubricant_type_id"],
            "running_hours": None,
            "photo_urls": [],
            "remarks": "Imported from legacy ACC master spreadsheet (LP_Change LOG).",
            "status": "approved",
            "submitted_at": d,
            "approved_by": None,
            "approved_at": d,
            "rejected_reason": None,
            "is_legacy_import": True,
        })
        rec_seq += 1

print("Historical lubrication_records:", len(lubrication_records))
last_change_dates = [s(cv(p["_row"], 'Last Change Date')) for p in lubrication_points]

# strip internal _row before writing
for p in lubrication_points:
    del p["_row"]

import os
os.makedirs('data/seed', exist_ok=True)
def dump(name, obj):
    with open(f'data/seed/{name}.json', 'w') as f:
        json.dump(obj, f, indent=2, default=str)

dump('organizations', organizations)
dump('areas', areas)
dump('equipment', equipment)
dump('lubricant_types', lubricant_types)
dump('lubrication_points', lubrication_points)
dump('lubrication_records', lubrication_records)
dump('data_quality_flags', data_quality_flags)

print("Distinct areas:", len(areas))
print("Distinct equipment:", len(equipment))
print("Distinct lubricant types:", len(lubricant_types))
print("Data quality flags logged:", len(data_quality_flags))
for f in data_quality_flags:
    print(" -", f)
